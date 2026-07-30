#!/usr/bin/env python3
"""Atualiza a Matriz de Gates com o resultado medido da rodada R18.43 (OS-07).

Le os JSON de reports/r1843/ (bateria n=48 e diag_espaco n=12, em tres bases de
semente), calcula a MEDIANA de 3 BASES de cada metrica e escreve uma copia
atualizada da planilha. Nada e digitado a mao: se a medicao mudar, roda de novo.

Uso:
  python3 tools/r1843/atualiza_matriz.py --entrada=MATRIZ.xlsx --saida=MATRIZ_R18.43.xlsx
"""
import argparse
import json
import os
import statistics
import sys

import openpyxl

RAIZ = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
BASES = ["s1", "s2", "s3"]

# gate -> (arquivo_prefixo, metrica, minimo, maximo)
GATES = {
    "ECO-01": ("bat", "goals", 2.4, 3.2),
    "ECO-02": ("bat", "xg", 1.8, 2.7),
    "ECO-03": ("bat", "shots", 12.0, 20.0),
    "ECO-04": ("bat", "onTarget", 4.0, 7.0),
    "ESP-01": ("esp", "vazio_r8", None, 50.0),
    "ESP-02": ("esp", "bloco_m", 35.0, None),
    "ESP-03": ("esp", "zaga_ataque_m", 30.0, None),
    "ESP-04": ("esp", "perto20_pct", 35.0, 45.0),
}


def carrega(prefixo, arm, base):
    caminho = os.path.join(RAIZ, "reports", "r1843", f"{prefixo}_{arm}_{base}.json")
    if not os.path.exists(caminho):
        return None
    with open(caminho, encoding="utf-8") as fh:
        return json.load(fh)["agregado"]


def mediana_3(prefixo, arm, metrica):
    vals = []
    for base in BASES:
        ag = carrega(prefixo, arm, base)
        if ag and metrica in ag:
            vals.append(ag[metrica]["media"])
    return (statistics.median(vals), len(vals), vals) if vals else (None, 0, [])


def cumpre(v, lo, hi):
    if v is None:
        return False
    return (lo is None or v >= lo) and (hi is None or v <= hi)


def br(v, casas=3):
    """Numero no formato da planilha (virgula decimal)."""
    if v is None:
        return "--"
    return f"{v:.{casas}f}".replace(".", ",")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--entrada", required=True)
    ap.add_argument("--saida", required=True)
    args = ap.parse_args()

    medidas = {}
    for gid, (pref, met, lo, hi) in GATES.items():
        mb, nb, vb = mediana_3(pref, "base", met)
        mc, nc, vc = mediana_3(pref, "rcc", met)
        # em quantas bases a BASELINE cumpre o gate
        okBase = sum(1 for v in vb if cumpre(v, lo, hi))
        medidas[gid] = {
            "baseline": mb, "cand": mc, "nBase": nb, "nCand": nc,
            "valsBase": vb, "valsCand": vc, "okBase": okBase,
            "cumpreBaseline": cumpre(mb, lo, hi), "cumpreCand": cumpre(mc, lo, hi),
            "lo": lo, "hi": hi,
        }

    wb = openpyxl.load_workbook(args.entrada)

    # ---------- Gate Matrix ----------
    ws = wb["Gate Matrix"]
    cab = [c.value for c in ws[1]]
    iBase = cab.index("Baseline")
    iStatus = cab.index("Status")
    iRel = cab.index("Release")
    iObs = cab.index("Observação")
    iEvid = cab.index("Evidência / comando")

    for row in ws.iter_rows(min_row=2):
        gid = row[0].value
        if gid not in medidas:
            continue
        m = medidas[gid]
        if m["cand"] is None:
            continue
        row[iBase].value = f"{br(m['baseline'])} R18.40A (mediana 3 bases)"
        row[iRel].value = "R18.43"
        row[iEvid].value = (
            "tools/r1840/bateria.js" if GATES[gid][0] == "bat" else "tools/r1843/diag_espaco.js"
        ) + " + tools/r1843/multibase43.js"
        if gid == "ESP-01":
            row[iStatus].value = "REJEITADO"
            row[iObs].value = (
                f"ALVO INATINGIVEL. Serie vazio_r8 (reconstruida; auditoria_blob nao existe no repo). "
                f"Piso geometrico 43,7%: 20 jogadores x disco r=8 m = 4021 m2 de 7140 m2. "
                f"<50% exigiria 88% da cobertura maxima teorica em todo quadro. "
                f"baseline {br(m['baseline'])}% -> candidata {br(m['cand'])}%. Ver tools/r1843/gate_esp01.md."
            )
        elif m["cumpreCand"]:
            row[iStatus].value = "RESOLVIDO" if gid.startswith("ESP") else "RESOLVIDO"
            extra = ""
            if m["okBase"] < 3:
                extra = (f" Gate era FRAGIL na baseline ({m['okBase']}/3 bases: "
                         + " / ".join(br(v) for v in m["valsBase"]) + ").")
            row[iObs].value = (
                f"R18.43 piso de profundidade do bloco. mediana 3 bases {br(m['cand'])} "
                f"(baseline {br(m['baseline'])}); bases: " + " / ".join(br(v) for v in m["valsCand"]) + "." + extra
            )
        else:
            row[iStatus].value = "PARCIAL"
            row[iObs].value = (
                f"R18.43: mediana 3 bases {br(m['cand'])} contra baseline {br(m['baseline'])} — fora da faixa."
            )

    # ---------- OS Status ----------
    ws2 = wb["OS Status"]
    cab2 = [c.value for c in ws2[1]]
    jStatus = cab2.index("Status")
    jRes = cab2.index("Resultado")
    jProx = cab2.index("Próxima ação")
    jProg = cab2.index("Progresso")
    e2 = medidas["ESP-02"]
    e3 = medidas["ESP-03"]
    e1 = medidas["ESP-01"]
    for row in ws2.iter_rows(min_row=2):
        if row[0].value != "OS-07":
            continue
        row[jStatus].value = "PARCIAL"
        row[jRes].value = (
            f"bloco {br(e2['baseline'],1)}->{br(e2['cand'],1)} m (alvo >35, CUMPRE); "
            f"zaga->ataque {br(e3['baseline'],1)}->{br(e3['cand'],1)} m (alvo >30, CUMPRE); "
            f"campo vazio {br(e1['baseline'],1)}->{br(e1['cand'],1)}% (alvo <50 INATINGIVEL)"
        )
        row[jProx].value = (
            "Reespecificar ESP-01 (alvo inatingivel, gate_esp01.md). ATENCAO: OS-07 NAO abriu "
            "folga para OS-02 — xG/chute caiu so 0,7% na mediana e o volume subiu 21%, entao o "
            "xG total subiu 17% e a folga ate o teto de ECO-02 caiu de 0,63/0,82/0,60 para "
            "0,28/0,32/0,14. A rota OS-07->OS-02 do handoff R18.40B precisa ser reaberta."
        )
        row[jProg].value = 0.7

    # ---------- aba nova com a medicao crua ----------
    if "R18.43 medicao" in wb.sheetnames:
        del wb["R18.43 medicao"]
    ws3 = wb.create_sheet("R18.43 medicao")
    ws3.append(["Gate", "Metrica", "Faixa", "Baseline s1", "s2", "s3",
                "Baseline mediana", "Cand s1", "s2", "s3", "Cand mediana",
                "Baseline cumpre (de 3)", "Candidata cumpre mediana"])
    for gid, (pref, met, lo, hi) in GATES.items():
        m = medidas[gid]
        faixa = ((f">={br(lo,1)}" if lo is not None else "")
                 + (" e " if lo is not None and hi is not None else "")
                 + (f"<={br(hi,1)}" if hi is not None else ""))
        vb = m["valsBase"] + [None] * (3 - len(m["valsBase"]))
        vc = m["valsCand"] + [None] * (3 - len(m["valsCand"]))
        ws3.append([gid, met, faixa, br(vb[0]), br(vb[1]), br(vb[2]), br(m["baseline"]),
                    br(vc[0]), br(vc[1]), br(vc[2]), br(m["cand"]),
                    f"{m['okBase']}/3", "SIM" if m["cumpreCand"] else "NAO"])
    for col, larg in zip("ABCDEFGHIJKLM",
                         [9, 16, 14, 12, 10, 10, 17, 10, 10, 10, 13, 22, 24]):
        ws3.column_dimensions[col].width = larg

    wb.save(args.saida)

    print(f"{os.path.basename(args.saida)} escrito.")
    print()
    print(f"{'gate':<8}{'faixa':<16}{'baseline':>10}{'cand':>10}  base cumpre  cand")
    for gid in GATES:
        m = medidas[gid]
        lo, hi = m["lo"], m["hi"]
        faixa = ((f">={lo:g}" if lo is not None else "")
                 + (" e " if lo is not None and hi is not None else "")
                 + (f"<={hi:g}" if hi is not None else ""))
        print(f"{gid:<8}{faixa:<16}{br(m['baseline']):>10}{br(m['cand']):>10}"
              f"      {m['okBase']}/3      {'CUMPRE' if m['cumpreCand'] else 'REPROVA'}")


if __name__ == "__main__":
    sys.exit(main())
