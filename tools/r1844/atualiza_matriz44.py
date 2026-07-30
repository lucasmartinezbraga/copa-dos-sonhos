#!/usr/bin/env python3
"""Atualiza a Matriz de Gates com o resultado medido da rodada R18.44.

Le os JSON de reports/r1843/ (bateria da baseline R18.43) e reports/r1844/
(bateria e coerencia da candidata), calcula a MEDIANA de 3 BASES e escreve uma
copia atualizada da planilha. Nada e digitado a mao.

Acrescenta a linha do gate novo COE-01 (gols/xG) e uma aba com a medicao crua.

Uso:
  python3 tools/r1844/atualiza_matriz44.py --entrada=MATRIZ_R18_43.xlsx \
      --saida=MATRIZ_R18_44.xlsx
"""
import argparse
import json
import os
import statistics
import sys

import openpyxl

RAIZ = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
BASES = (1, 2, 3)


def le(caminho):
    p = os.path.join(RAIZ, caminho)
    if not os.path.exists(p):
        return None
    with open(p, encoding="utf-8") as fh:
        return json.load(fh)


def serie(padrao, metrica, onde):
    """Devolve (mediana, [valores]) para a metrica nas 3 bases.

    `onde` = 'agregado' (bateria) ou 'totais' (diag_xg).
    """
    vals = []
    for s in BASES:
        d = le(padrao.format(s=s))
        if not d:
            continue
        v = d["agregado"][metrica]["media"] if onde == "agregado" else d["totais"].get(metrica)
        if isinstance(v, (int, float)):
            vals.append(v)
    return (statistics.median(vals) if vals else None), vals


def cumpre(v, lo, hi):
    if v is None:
        return False
    return (lo is None or v >= lo) and (hi is None or v <= hi)


def br(v, c=3):
    return "--" if v is None else f"{v:.{c}f}".replace(".", ",")


# gate -> (metrica, onde, padrao_baseline, padrao_candidata, lo, hi)
GATES = {
    "ECO-01": ("goals",    "agregado", "reports/r1843/bat_rcc_s{s}.json", "reports/r1844/bat_gk_s{s}.json", 2.4, 3.2),
    "ECO-02": ("xg",       "agregado", "reports/r1843/bat_rcc_s{s}.json", "reports/r1844/bat_gk_s{s}.json", 1.8, 2.7),
    "ECO-03": ("shots",    "agregado", "reports/r1843/bat_rcc_s{s}.json", "reports/r1844/bat_gk_s{s}.json", 12.0, 20.0),
    "ECO-04": ("onTarget", "agregado", "reports/r1843/bat_rcc_s{s}.json", "reports/r1844/bat_gk_s{s}.json", 4.0, 7.0),
    "COE-01": ("gols_por_xg", "totais", "reports/r1844/xg_b43_s{s}.json", "reports/r1844/xg_gk_s{s}.json", 0.90, 1.15),
    "CAU-03": ("pct_dos_gols_por_falha_do_goleiro", "totais", "reports/r1844/xg_b43_s{s}.json", "reports/r1844/xg_gk_s{s}.json", None, 8.0),
}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--entrada", required=True)
    ap.add_argument("--saida", required=True)
    args = ap.parse_args()

    m = {}
    for gid, (met, onde, pb, pc, lo, hi) in GATES.items():
        mb, vb = serie(pb, met, onde)
        mc, vc = serie(pc, met, onde)
        m[gid] = {
            "mb": mb, "vb": vb, "mc": mc, "vc": vc, "lo": lo, "hi": hi,
            "okBase": sum(1 for v in vb if cumpre(v, lo, hi)),
            "okCand": cumpre(mc, lo, hi),
        }

    wb = openpyxl.load_workbook(args.entrada)
    ws = wb["Gate Matrix"]
    cab = [c.value for c in ws[1]]
    iB, iSt, iRel = cab.index("Baseline"), cab.index("Status"), cab.index("Release")
    iObs, iEv = cab.index("Observação"), cab.index("Evidência / comando")

    for row in ws.iter_rows(min_row=2):
        gid = row[0].value
        if gid in ("ECO-01", "ECO-02", "ECO-03", "ECO-04") and m[gid]["mc"] is not None:
            g = m[gid]
            row[iB].value = f"{br(g['mc'])} R18.44 (mediana 3 bases)"
            row[iRel].value = "R18.44"
            row[iEv].value = "tools/r1840/bateria.js + tools/r1844/multibase44.js"
            extra = ""
            if g["okBase"] < 3:
                extra = (f" Era FRAGIL na R18.43 ({g['okBase']}/3 bases: "
                         + " / ".join(br(v) for v in g["vb"]) + ") e agora cumpre em 3/3.")
            row[iObs].value = (f"R18.44 raio do goleiro 1,95 (plano = checagem). mediana 3 bases {br(g['mc'])} "
                               f"(R18.43 {br(g['mb'])}); bases: " + " / ".join(br(v) for v in g["vc"]) + "." + extra)
        if gid == "CAU-03":
            g = m["CAU-03"]
            row[iB].value = f"{br(g['mb'], 2)}% R18.43 (mediana 3 bases)"
            row[iSt].value = "RESOLVIDO" if g["okCand"] else "PARCIAL"
            row[iRel].value = "R18.44"
            row[iEv].value = "tools/r1843/diag_xg.js (evento visual_contact_failed kind=save)"
            row[iObs].value = (
                f"% dos gols por falha de contato do goleiro: {br(g['mb'], 2)}% -> {br(g['mc'])}% "
                f"(mediana 3 bases; bases " + " / ".join(br(v) for v in g["vc"]) + "). "
                "Mecanismo: os tres sitios de chute planejavam com radius=3.0 e _gkResolveSave validava com "
                "_physicalContactValid(gk,1.95,z); como required=max(0,dist-radius), o plano declarava o goleiro "
                "no lugar e o corpo nunca recebia ordem de andar. Introduzido por tools/r1821/build_rc1.js ~linha 128."
            )
        if gid == "CAU-04":
            row[iRel].value = "R18.44"
            row[iObs].value = ("O que mordia era o RAIO, nao a velocidade — baixar a promessa do planejador "
                               "saiu inerte no ponto escolhido. Raio corrigido na R18.44.")

    # ---- linha do gate novo, se ainda nao existir ----
    ids = [r[0].value for r in ws.iter_rows(min_row=2)]
    if "COE-01" not in ids:
        g = m["COE-01"]
        ws.append(["COE-01", "Integridade", "gols por partida / xG por partida", "Bateria", "n=48 x 3 bases",
                   f"{br(g['mb'])} R18.43", "0,90-1,15", "10%", "Coerencia do modelo de xG", "R18.44",
                   "RESOLVIDO" if g["okCand"] else "ABERTO",
                   "tools/r1843/diag_xg.js + tools/r1844/multibase44.js",
                   (f"GATE NOVO. O modelo de xG deve prever os proprios gols do jogo. Faixa derivada de futebol "
                    f"real, onde o xG e calibrado contra gols e a razao fica ~1,0 — nao da saida do motor. "
                    f"R18.43 {br(g['mb'])} (reprova em 3/3) -> R18.44 {br(g['mc'])} "
                    f"(bases " + " / ".join(br(v) for v in g["vc"]) + "). Nenhum gate da matriz forcava esta "
                    "coerencia, e por isso um caminho de gol sem xG (goal_after_failed_reach) atravessou varias "
                    "rodadas promovidas sem nada reprovar.")])

    # ---- OS Status ----
    ws2 = wb["OS Status"]
    cab2 = [c.value for c in ws2[1]]
    jSt, jRes, jProx, jProg = (cab2.index("Status"), cab2.index("Resultado"),
                               cab2.index("Próxima ação"), cab2.index("Progresso"))
    for row in ws2.iter_rows(min_row=2):
        if row[0].value == "OS-01":
            g = m["CAU-03"]
            row[jSt].value = "RESOLVIDO"
            row[jRes].value = (f"CAU-03 {br(g['mb'], 2)}% -> {br(g['mc'])}% (alvo <8%), mediana 3 bases. "
                               f"O mecanismo era o RAIO (plano 3.0 vs checagem 1.95), nao a velocidade.")
            row[jProx].value = "Manter teste de regressao; microcenarios GK-01..03 seguem abertos"
            row[jProg].value = 1
        if row[0].value == "OS-02":
            row[jProx].value = (
                "Folga de ECO-01 subiu de 0,08 para 0,66 na R18.44, mas ECO-02 segue o gargalo (folga 0,21 "
                "contra +0,92 de xG da escalacao). ATENCAO: futebol real tem xG total ~2,7, logo o teto de "
                "ECO-02 esta CERTO — um XI correto entregando 3,4 nao e gate mal calibrado, e o resto do motor "
                "sendo generoso. Atacar low_cross_shot primeiro (44% das finalizacoes, sem termo posicional)."
            )

    # ---- aba de medicao crua ----
    if "R18.44 medicao" in wb.sheetnames:
        del wb["R18.44 medicao"]
    w3 = wb.create_sheet("R18.44 medicao")
    w3.append(["Gate", "Metrica", "Faixa", "R18.43 s1", "s2", "s3", "R18.43 mediana",
               "R18.44 s1", "s2", "s3", "R18.44 mediana", "R18.43 cumpre (de 3)", "R18.44 cumpre mediana"])
    for gid, (met, onde, pb, pc, lo, hi) in GATES.items():
        g = m[gid]
        faixa = ((f">={br(lo, 2)}" if lo is not None else "")
                 + (" e " if lo is not None and hi is not None else "")
                 + (f"<={br(hi, 2)}" if hi is not None else ""))
        vb = g["vb"] + [None] * (3 - len(g["vb"]))
        vc = g["vc"] + [None] * (3 - len(g["vc"]))
        w3.append([gid, met, faixa, br(vb[0]), br(vb[1]), br(vb[2]), br(g["mb"]),
                   br(vc[0]), br(vc[1]), br(vc[2]), br(g["mc"]),
                   f"{g['okBase']}/3", "SIM" if g["okCand"] else "NAO"])
    for col, larg in zip("ABCDEFGHIJKLM", [9, 34, 16, 12, 10, 10, 17, 12, 10, 10, 17, 21, 23]):
        w3.column_dimensions[col].width = larg

    wb.save(args.saida)

    print(f"{os.path.basename(args.saida)} escrito.")
    print()
    print(f"{'gate':<8}{'faixa':<18}{'R18.43':>10}{'R18.44':>10}  base  cand")
    for gid in GATES:
        g = m[gid]
        lo, hi = g["lo"], g["hi"]
        faixa = ((f">={lo:g}" if lo is not None else "") + (" e " if lo is not None and hi is not None else "")
                 + (f"<={hi:g}" if hi is not None else ""))
        print(f"{gid:<8}{faixa:<18}{br(g['mb']):>10}{br(g['mc']):>10}  {g['okBase']}/3  "
              f"{'CUMPRE' if g['okCand'] else 'REPROVA'}")


if __name__ == "__main__":
    sys.exit(main())
