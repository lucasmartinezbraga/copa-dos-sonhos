#!/usr/bin/env python3
"""Registra na Matriz de Gates a re-derivacao de ECO-03/ECO-04 APROVADA pelo dono
da matriz, e reavalia as builds da linhagem sob as faixas novas.

A decisao aprovada (ver tools/r1845/gate_eco_consistencia.md):

    ECO-03  chutes por partida   17 - 27   (era 12 - 20)
    ECO-04  chutes no alvo        6 - 10   (era 4 - 7)

derivadas de futebol real (~25 chutes e ~8,5 no alvo, somando os dois times) na
faixa 0,7-1,0x que a Ordem de Servico define para camadas de "Momento".
ECO-01 (2,4-3,2 gols) e ECO-02 (1,8-2,7 xG) ficam como estao.

CONSEQUENCIA QUE PRECISA ESTAR ESCRITA NA PLANILHA: sob as faixas novas, TODAS as
builds da linhagem — inclusive a R18.44 promovida — reprovam ECO-03 e ECO-04. Isso
nao e efeito colateral indesejado: e o proposito. As faixas antigas escondiam que
o motor chuta a ~58% do volume do futebol real. Pela regra estabelecida na R18.40
(um gate que a propria baseline reprova nao bloqueia promocao e vira defeito
rastreado), ECO-03 e ECO-04 passam a ser DEFEITO RASTREADO com alvo claro, e nao
bloqueio.

Uso:
  python3 tools/r1846/aplica_decisao_gates.py --entrada=MATRIZ.xlsx --saida=SAIDA.xlsx
"""
import argparse
import json
import os
import statistics
import sys

import openpyxl

RAIZ = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

NOVAS = {"ECO-03": (17.0, 27.0, "17-27", "~25 (dois times)"),
         "ECO-04": (6.0, 10.0, "6-10", "~8,5 (dois times)")}
ANTIGAS = {"ECO-03": "12-20", "ECO-04": "4-7"}
METRICA = {"ECO-03": "shots", "ECO-04": "onTarget"}

BUILDS = [
    ("R18.43", "reports/r1843/bat_rcc_s{s}.json"),
    ("R18.44", "reports/r1844/bat_gk_s{s}.json"),
    ("R18.45-RC b2,60", "reports/r1845/bat_lc2.60_s{s}.json"),
    ("R18.45-RC b1,35", "reports/r1845/bat_lc1.35_s{s}.json"),
]


def serie(pat, met):
    vals = []
    for s in (1, 2, 3):
        p = os.path.join(RAIZ, pat.format(s=s))
        if os.path.exists(p):
            with open(p, encoding="utf-8") as fh:
                vals.append(json.load(fh)["agregado"][met]["media"])
    return (statistics.median(vals) if vals else None), vals


def br(v, c=3):
    return "--" if v is None else f"{v:.{c}f}".replace(".", ",")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--entrada", required=True)
    ap.add_argument("--saida", required=True)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.entrada)
    ws = wb["Gate Matrix"]
    cab = [c.value for c in ws[1]]
    iAlvo, iSt, iObs, iRel = (cab.index("Alvo"), cab.index("Status"),
                              cab.index("Observação"), cab.index("Release"))
    iBase = cab.index("Baseline")

    # R18.44 e a build promovida: e a referencia para dizer se o gate e valido
    ref = {}
    for gid in NOVAS:
        ref[gid] = serie("reports/r1844/bat_gk_s{s}.json", METRICA[gid])

    for row in ws.iter_rows(min_row=2):
        gid = row[0].value
        if gid not in NOVAS:
            continue
        lo, hi, txt, real = NOVAS[gid]
        med, vals = ref[gid]
        okN = sum(1 for v in vals if lo <= v <= hi)
        row[iAlvo].value = txt
        row[iRel].value = "R18.46"
        row[iSt].value = "ABERTO" if okN == 0 else "PARCIAL"
        row[iBase].value = f"{br(med)} R18.44 (mediana 3 bases)"
        row[iObs].value = (
            f"FAIXA RE-DERIVADA E APROVADA pelo dono da matriz (era {ANTIGAS[gid]}). "
            f"Derivada de futebol real {real} na faixa 0,7-1,0x da Ordem de Servico para camadas "
            f"de 'Momento'. Motivo: a faixa antiga EXCLUIA o futebol real (teto de 20 chutes contra "
            f"~25 reais) e por isso escondia que o motor chuta a ~58% do volume real. "
            f"Sob a faixa nova a build promovida R18.44 entrega {br(med)} e REPROVA em {okN}/3 bases "
            f"({' / '.join(br(v) for v in vals)}) — pela regra da R18.40, gate que a baseline reprova "
            f"nao bloqueia promocao e vira DEFEITO RASTREADO com alvo claro. "
            f"Prova de insatisfazibilidade do conjunto antigo em tools/r1845/gate_eco_consistencia.md."
        )

    # aba com a reavaliacao de todas as builds sob as faixas novas
    nome = "R18.46 faixas novas"
    if nome in wb.sheetnames:
        del wb[nome]
    w = wb.create_sheet(nome)
    w.append(["DECISAO APROVADA: ECO-03 17-27 (era 12-20) e ECO-04 6-10 (era 4-7)"])
    w.append(["Derivadas de futebol real na faixa 0,7-1,0x da Ordem de Servico. ECO-01 e ECO-02 inalterados."])
    w.append([])
    w.append(["build", "gate", "s1", "s2", "s3", "mediana",
              "faixa antiga", "cumpria antiga", "faixa nova", "cumpre nova"])
    ANT = {"ECO-03": (12.0, 20.0), "ECO-04": (4.0, 7.0)}
    for nomeB, pat in BUILDS:
        for gid in ("ECO-03", "ECO-04"):
            med, vals = serie(pat, METRICA[gid])
            if med is None:
                continue
            lo, hi, txt, _ = NOVAS[gid]
            alo, ahi = ANT[gid]
            v = vals + [None] * (3 - len(vals))
            w.append([nomeB, gid, br(v[0]), br(v[1]), br(v[2]), br(med),
                      ANTIGAS[gid], "SIM" if alo <= med <= ahi else "NAO",
                      txt, "SIM" if lo <= med <= hi else "NAO"])
    w.append([])
    w.append(["futebol real (soma dos dois times)", "", "", "", "", "chutes ~25 · no alvo ~8,5 · gols ~2,7 · xG ~2,7"])
    for col, larg in zip("ABCDEFGHIJ", [18, 9, 10, 10, 10, 10, 14, 16, 12, 13]):
        w.column_dimensions[col].width = larg

    wb.save(args.saida)
    print(f"{os.path.basename(args.saida)} escrito.")
    print()
    print(f"{'build':<18}{'gate':<9}{'mediana':>9}{'antiga':>9}{'nova':>7}")
    for nomeB, pat in BUILDS:
        for gid in ("ECO-03", "ECO-04"):
            med, vals = serie(pat, METRICA[gid])
            if med is None:
                continue
            lo, hi, _, _ = NOVAS[gid]
            alo, ahi = ANT[gid]
            print(f"{nomeB:<18}{gid:<9}{br(med):>9}"
                  f"{('ok' if alo <= med <= ahi else 'REP'):>9}{('ok' if lo <= med <= hi else 'REP'):>7}")


if __name__ == "__main__":
    sys.exit(main())
